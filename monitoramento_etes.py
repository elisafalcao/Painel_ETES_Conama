"""
Monitoramento de ETEs - CONAMA 430/2011

Gera tabela consolidada, status de conformidade, planilha Excel formatada,
HTML e gráficos a partir de uma planilha de resultados laboratoriais.

Uso:
    python monitoramento_etes.py --entrada ResultadoETEsCESAN_JAN2025python.xlsx --saida resultados
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt


# =========================
# Parâmetros legais adotados
# =========================
# CONAMA 430/2011, Art. 16 e Art. 21, para lançamento de efluentes sanitários.
# Observação: alguns órgãos ambientais/licenças podem estabelecer limites mais restritivos.
LIMITES = {
    "ph_min": 5.0,
    "ph_max": 9.0,
    "temperatura_max_c": 40.0,
    "materiais_sedimentaveis_max_ml_l": 1.0,
    "dbo_eficiencia_min_pct": 60.0,
    # Para esgoto sanitário, usa-se 50 mg/L como referência para óleos vegetais e gorduras animais.
    # Caso a licença exija 20 mg/L ou outro limite, altere este valor.
    "oleos_graxas_max_mg_l": 50.0,
    "materiais_flutuantes_ok": "ausente",
}

CORES = {
    "Conforme": "#C6EFCE",
    "Atenção": "#FFEB9C",
    "Não conforme": "#FFC7CE",
    "Sem dado": "#D9EAF7",
}

PALETA = {
    "azul": "#5DADE2",
    "azul_escuro": "#2874A6",
    "rosa": "#F5B7B1",
    "rosa_escuro": "#C0392B",
    "cinza": "#7F8C8D",
    "cinza_claro": "#F4F6F7",
}

COLUNAS_FINAIS = [
    "Mês / Ano",
    "Estação",
    "pH Efluente",
    "Temperatura Efluente",
    "Materiais Sedimentares Efluente",
    "DBO Afluente",
    "DBO Efluente",
    "Eficiência DBO",
    "Óleos e Graxas Efluente",
    "Materiais Flutuantes Efluente",
]


def normalizar_texto(valor: Any) -> str:
    if pd.isna(valor):
        return ""
    return str(valor).strip()


def para_numero(valor: Any) -> float:
    if pd.isna(valor) or valor == "":
        return np.nan
    texto = str(valor).replace(",", ".")
    texto = re.sub(r"[^0-9.\-]", "", texto)
    try:
        return float(texto)
    except ValueError:
        return np.nan


def ler_dados(caminho: Path) -> pd.DataFrame:
    df = pd.read_excel(caminho, sheet_name=0)
    df.columns = [normalizar_texto(c).replace("\n", " ") for c in df.columns]

    renomear = {
        "Ident. do Componente": "Estação",
        "PONTO DE COLETA": "Ponto",
        "Data": "Data",
        "DBO mg O2/L": "DBO",
        "Mat.Flut": "Materiais Flutuantes",
        "O&G mg/L": "Óleos e Graxas",
        "pH": "pH",
        "SSed mL/L": "Materiais Sedimentares",
        "Temp(am) ºC": "Temperatura",
    }
    df = df.rename(columns=renomear)

    obrigatorias = ["Estação", "Ponto", "Data", "DBO", "Materiais Flutuantes", "Óleos e Graxas", "pH", "Materiais Sedimentares", "Temperatura"]
    faltantes = [c for c in obrigatorias if c not in df.columns]
    if faltantes:
        raise ValueError(f"Colunas obrigatórias não encontradas: {faltantes}")

    df["Estação"] = df["Estação"].map(normalizar_texto)
    df["Ponto"] = df["Ponto"].map(lambda x: normalizar_texto(x).lower())
    df["Data"] = pd.to_datetime(df["Data"], dayfirst=True, errors="coerce")
    df["Mês / Ano"] = df["Data"].dt.strftime("%m/%Y")

    for col in ["DBO", "Óleos e Graxas", "pH", "Materiais Sedimentares", "Temperatura"]:
        df[col] = df[col].map(para_numero)

    return df


def consolidar_tabela(df: pd.DataFrame) -> pd.DataFrame:
    afluente = df[df["Ponto"].str.contains("afluente", na=False)].copy()
    efluente = df[df["Ponto"].str.contains("efluente", na=False)].copy()

    chaves = ["Estação", "Mês / Ano"]
    agg_num = "mean"

    af = afluente.groupby(chaves, as_index=False).agg({"DBO": agg_num}).rename(columns={"DBO": "DBO Afluente"})
    ef = efluente.groupby(chaves, as_index=False).agg({
        "pH": agg_num,
        "Temperatura": agg_num,
        "Materiais Sedimentares": agg_num,
        "DBO": agg_num,
        "Óleos e Graxas": agg_num,
        "Materiais Flutuantes": lambda s: "; ".join(sorted(set(normalizar_texto(v) for v in s if normalizar_texto(v))))
    }).rename(columns={
        "pH": "pH Efluente",
        "Temperatura": "Temperatura Efluente",
        "Materiais Sedimentares": "Materiais Sedimentares Efluente",
        "DBO": "DBO Efluente",
        "Óleos e Graxas": "Óleos e Graxas Efluente",
        "Materiais Flutuantes": "Materiais Flutuantes Efluente",
    })

    tabela = ef.merge(af, on=chaves, how="left")
    tabela["Eficiência DBO"] = (1 - tabela["DBO Efluente"] / tabela["DBO Afluente"]) * 100
    tabela = tabela[COLUNAS_FINAIS].sort_values(["Mês / Ano", "Estação"]).reset_index(drop=True)
    return tabela


def status_intervalo(valor: float, minimo: float | None = None, maximo: float | None = None, margem_pct: float = 0.10) -> str:
    if pd.isna(valor):
        return "Sem dado"
    if minimo is not None and valor < minimo:
        return "Não conforme"
    if maximo is not None and valor > maximo:
        return "Não conforme"

    # Amarelo quando o valor está próximo do limite.
    if minimo is not None and valor <= minimo * (1 + margem_pct):
        return "Atenção"
    if maximo is not None and valor >= maximo * (1 - margem_pct):
        return "Atenção"
    return "Conforme"


def status_texto_ausente(valor: Any) -> str:
    txt = normalizar_texto(valor).lower()
    if not txt:
        return "Sem dado"
    return "Conforme" if LIMITES["materiais_flutuantes_ok"] in txt else "Não conforme"


def adicionar_status(tabela: pd.DataFrame) -> pd.DataFrame:
    out = tabela.copy()
    out["Status pH"] = out["pH Efluente"].map(lambda x: status_intervalo(x, LIMITES["ph_min"], LIMITES["ph_max"]))
    out["Status Temperatura"] = out["Temperatura Efluente"].map(lambda x: status_intervalo(x, None, LIMITES["temperatura_max_c"]))
    out["Status Materiais Sedimentares"] = out["Materiais Sedimentares Efluente"].map(lambda x: status_intervalo(x, None, LIMITES["materiais_sedimentaveis_max_ml_l"]))
    out["Status Eficiência DBO"] = out["Eficiência DBO"].map(lambda x: status_intervalo(x, LIMITES["dbo_eficiencia_min_pct"], None))
    out["Status Óleos e Graxas"] = out["Óleos e Graxas Efluente"].map(lambda x: status_intervalo(x, None, LIMITES["oleos_graxas_max_mg_l"]))
    out["Status Materiais Flutuantes"] = out["Materiais Flutuantes Efluente"].map(status_texto_ausente)

    status_cols = [c for c in out.columns if c.startswith("Status ")]
    out["Itens conformes"] = (out[status_cols] == "Conforme").sum(axis=1)
    out["Itens não conformes"] = (out[status_cols] == "Não conforme").sum(axis=1)
    out["Itens em atenção"] = (out[status_cols] == "Atenção").sum(axis=1)
    out["Itens sem dado"] = (out[status_cols] == "Sem dado").sum(axis=1)
    out["Índice de Conformidade (%)"] = (out["Itens conformes"] / len(status_cols) * 100).round(1)

    def status_geral(row: pd.Series) -> str:
        if row["Itens não conformes"] > 0:
            return "Não conforme"
        if row["Itens em atenção"] > 0 or row["Itens sem dado"] > 0:
            return "Atenção"
        return "Conforme"

    out["Status Geral"] = out.apply(status_geral, axis=1)
    return out


def salvar_excel(tabela: pd.DataFrame, saida: Path) -> Path:
    saida.mkdir(parents=True, exist_ok=True)
    caminho = saida / "tabela_conformidade_etes.xlsx"

    with pd.ExcelWriter(caminho, engine="openpyxl") as writer:
        tabela.to_excel(writer, index=False, sheet_name="Conformidade")
        limites_df = pd.DataFrame([LIMITES]).T.reset_index()
        limites_df.columns = ["Parâmetro", "Limite adotado"]
        limites_df.to_excel(writer, index=False, sheet_name="Limites")

        wb = writer.book
        ws = wb["Conformidade"]
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.formatting.rule import FormulaRule

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True, color="1F4E79")
        thin = Side(style="thin", color="D9D9D9")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)

        widths = {
            "A": 12, "B": 32, "C": 12, "D": 18, "E": 24, "F": 14, "G": 14, "H": 15,
            "I": 22, "J": 24
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
        for col in range(11, ws.max_column + 1):
            ws.column_dimensions[ws.cell(1, col).column_letter].width = 18

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = Border(bottom=thin)

        percent_cols = ["H", ws.cell(1, tabela.columns.get_loc("Índice de Conformidade (%)") + 1).column_letter]
        for col in percent_cols:
            for cell in ws[col][1:]:
                cell.number_format = "0.0"

        # Cores para células de status.
        for col_idx, col_name in enumerate(tabela.columns, start=1):
            if col_name.startswith("Status"):
                col_letter = ws.cell(1, col_idx).column_letter
                rng = f"{col_letter}2:{col_letter}{ws.max_row}"
                for status, cor in CORES.items():
                    ws.conditional_formatting.add(
                        rng,
                        FormulaRule(formula=[f'{col_letter}2="{status}"'], fill=PatternFill("solid", fgColor=cor.replace("#", "")))
                    )

    return caminho


def salvar_html(tabela: pd.DataFrame, saida: Path) -> Path:
    caminho = saida / "tabela_conformidade_etes.html"

    def colorir_status(valor: Any) -> str:
        return f"background-color: {CORES.get(str(valor), 'white')}"

    styled = (
        tabela.style
        .map(colorir_status, subset=[c for c in tabela.columns if c.startswith("Status") or c == "Status Geral"])
        .format({
            "pH Efluente": "{:.2f}",
            "Temperatura Efluente": "{:.1f}",
            "Materiais Sedimentares Efluente": "{:.2f}",
            "DBO Afluente": "{:.1f}",
            "DBO Efluente": "{:.1f}",
            "Eficiência DBO": "{:.1f}%",
            "Óleos e Graxas Efluente": "{:.2f}",
            "Índice de Conformidade (%)": "{:.1f}%",
        }, na_rep="-")
        .set_table_styles([
            {"selector": "th", "props": [("background-color", PALETA["azul"]), ("color", "white"), ("text-align", "center")]},
            {"selector": "td", "props": [("font-family", "Arial"), ("font-size", "12px")]},
            {"selector": "caption", "props": [("caption-side", "top"), ("font-size", "20px"), ("font-weight", "bold"), ("color", PALETA["azul_escuro"])]},
        ])
        .set_caption("Monitoramento de ETEs - Conformidade CONAMA 430/2011")
    )
    caminho.write_text(styled.to_html(), encoding="utf-8")
    return caminho


def salvar_graficos(tabela: pd.DataFrame, saida: Path) -> list[Path]:
    graficos_dir = saida / "graficos"
    graficos_dir.mkdir(parents=True, exist_ok=True)
    arquivos = []

    plt.rcParams.update({
        "figure.facecolor": "white",
        "axes.facecolor": PALETA["cinza_claro"],
        "axes.edgecolor": "#D5D8DC",
        "font.size": 10,
    })

    # 1) Distribuição de status geral
    contagem = tabela["Status Geral"].value_counts().reindex(["Conforme", "Atenção", "Não conforme"], fill_value=0)
    fig, ax = plt.subplots(figsize=(8, 5))
    ax.bar(contagem.index, contagem.values, color=["#82E0AA", "#F7DC6F", "#F1948A"])
    ax.set_title("Distribuição do Status Geral das ETEs")
    ax.set_ylabel("Quantidade")
    for i, v in enumerate(contagem.values):
        ax.text(i, v, str(v), ha="center", va="bottom")
    fig.tight_layout()
    p = graficos_dir / "01_distribuicao_status_geral.png"
    fig.savefig(p, dpi=180)
    plt.close(fig)
    arquivos.append(p)

    # 2) Top 10 conformidade
    top_conf = tabela.sort_values(["Índice de Conformidade (%)", "Itens não conformes"], ascending=[False, True]).head(10)
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.barh(top_conf["Estação"], top_conf["Índice de Conformidade (%)"], color=PALETA["azul"])
    ax.invert_yaxis()
    ax.set_xlim(0, 100)
    ax.set_title("Top 10 Estações com Maior Conformidade")
    ax.set_xlabel("Índice de Conformidade (%)")
    fig.tight_layout()
    p = graficos_dir / "02_top10_maior_conformidade.png"
    fig.savefig(p, dpi=180)
    plt.close(fig)
    arquivos.append(p)

    # 3) Top 10 não conformidade
    top_nc = tabela.sort_values(["Itens não conformes", "Índice de Conformidade (%)"], ascending=[False, True]).head(10)
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.barh(top_nc["Estação"], top_nc["Itens não conformes"], color=PALETA["rosa"])
    ax.invert_yaxis()
    ax.set_title("Top 10 Estações com Maior Não Conformidade")
    ax.set_xlabel("Número de parâmetros não conformes")
    fig.tight_layout()
    p = graficos_dir / "03_top10_maior_nao_conformidade.png"
    fig.savefig(p, dpi=180)
    plt.close(fig)
    arquivos.append(p)

    # 4) Eficiência DBO por estação
    dados = tabela.sort_values("Eficiência DBO", ascending=True).head(20)
    fig, ax = plt.subplots(figsize=(11, 7))
    ax.barh(dados["Estação"], dados["Eficiência DBO"], color=PALETA["azul_escuro"])
    ax.axvline(LIMITES["dbo_eficiencia_min_pct"], linestyle="--", color=PALETA["rosa_escuro"], label="Mínimo CONAMA 430/2011: 60%")
    ax.set_title("20 Menores Eficiências de Remoção de DBO")
    ax.set_xlabel("Eficiência DBO (%)")
    ax.legend()
    fig.tight_layout()
    p = graficos_dir / "04_menores_eficiencias_dbo.png"
    fig.savefig(p, dpi=180)
    plt.close(fig)
    arquivos.append(p)

    return arquivos


def main() -> None:
    parser = argparse.ArgumentParser(description="Gera tabela e gráficos de conformidade de ETEs.")
    parser.add_argument("--entrada", required=True, help="Caminho da planilha .xlsx de entrada")
    parser.add_argument("--saida", default="resultados", help="Pasta de saída")
    args = parser.parse_args()

    entrada = Path(args.entrada)
    saida = Path(args.saida)

    df = ler_dados(entrada)
    tabela = consolidar_tabela(df)
    tabela = adicionar_status(tabela)

    excel = salvar_excel(tabela, saida)
    html = salvar_html(tabela, saida)
    graficos = salvar_graficos(tabela, saida)

    print("Arquivos gerados:")
    print(f"- {excel}")
    print(f"- {html}")
    for g in graficos:
        print(f"- {g}")


if __name__ == "__main__":
    main()
